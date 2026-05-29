"""
eBay 页面爬虫 - Playwright 浏览器自动化搜索
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
直接爬取 eBay 搜索结果页面，无需 API Key

特点:
- 使用 Playwright + playwright-stealth 反检测
- 适配 eBay 2025 新版 DOM 结构 (s-card)
- 支持 BestMatch / 价格升序 / 价格+运费升序 等排序
- 自动设置美国 locale，确保搜到美国站结果
- 提取: 标题、价格、运费、图片URL、商品URL、商品ID、状况等
- 输出结构化 JSON / Excel，兼容 competitor_finder.py

用法:
    python ebay_page_scraper.py "Honda Pilot running boards"
    python ebay_page_scraper.py "关键词" --sort best_match --last 10 --pages 2
    python ebay_page_scraper.py "关键词" --pages 0 --excel
"""

import re
import io
import sys
import os
import json
import time
import random
import argparse
from pathlib import Path

from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

# ── Windows 控制台编码修复 ──
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ── CNY → USD 汇率（固定值） ──
_CNY_USD_RATE = 7.836


def get_cny_usd_rate() -> float:
    """返回 CNY/USD 汇率"""
    return _CNY_USD_RATE

# 默认输出文件
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUTPUT = os.path.join(_SCRIPT_DIR, "_tmp_ebay_page_result.json")

# ── 排序参数映射 ──
SORT_MAP = {
    "best_match": None,
    "price_low": "_sop=2",
    "price_high": "_sop=3",
    "price_ship": "_sop=15",
    "newest": "_sop=10",
    "ending": "_sop=1",
}

SORT_DESC = {
    "best_match": "Best Match",
    "price_low": "Price: Low to High",
    "price_high": "Price: High to Low",
    "price_ship": "Price + Shipping: Low to High",
    "newest": "Newly Listed",
    "ending": "Ending Soonest",
}


def build_url(keyword: str, sort: str = "best_match", page: int = 1, marketplace: str = "EBAY-US") -> str:
    """构建 eBay 搜索 URL"""
    from urllib.parse import quote_plus
    k = quote_plus(keyword)
    # 根据市场选择对应货币
    currency_map = {
        "EBAY-US": "USD",
        "EBAY-GB": "GBP",
        "EBAY-DE": "EUR",
        "EBAY-AU": "AUD",
        "EBAY-CA": "CAD",
    }
    currency = currency_map.get(marketplace, "USD")
    url = f"https://www.ebay.com/sch/i.html?_nkw={k}&_sacat=0&_from=R40&currency={currency}"

    sop = SORT_MAP.get(sort)
    if sop:
        url += f"&{sop}"

    if page > 1:
        url += f"&_pgn={page}"

    return url


# ── JS 提取脚本 ──
JS_EXTRACT = r"""() => {
    let totalResults = 0;
    const countEl = document.querySelector('.srp-controls__count-heading, h1.srp-controls__count-heading');
    if (countEl) {
        const cm = countEl.textContent.replace(/,/g, '').match(/(\d+)/);
        if (cm) totalResults = parseInt(cm[1]);
    }
    if (!totalResults) {
        const pageLinks = document.querySelectorAll('a.pagination__item');
        if (pageLinks.length) {
            const lastPage = pageLinks[pageLinks.length - 1];
            const pm = lastPage.textContent.trim().match(/(\d+)/);
            if (pm) totalResults = parseInt(pm[1]) * 60;
        }
    }

    const CNY_RATE = window.__CNY_RATE__ || 6.8;
    const FORCE_CNY = window.__FORCE_CNY__ || false;

    const results = [];
    let cards = document.querySelectorAll('li.s-card');
    if (cards.length === 0) cards = document.querySelectorAll('li.s-item');

    cards.forEach(card => {
        let title = '';
        const titleEl = card.querySelector('.s-card__title') || card.querySelector('.s-item__title');
        if (titleEl) {
            title = titleEl.textContent.trim();
        }
        if (!title || title === 'Shop on eBay' || title.startsWith('Click to')) return;

        let itemUrl = '';
        let itemId = '';
        const linkEl = card.querySelector('a.s-card__link[href*="/itm/"]') || card.querySelector('a[href*="/itm/"]');
        if (linkEl) {
            itemUrl = linkEl.getAttribute('href') || '';
            const m = itemUrl.match(/\/itm\/(\d+)/);
            if (m) itemId = m[1];
        }

        let price = null;
        const priceEl = card.querySelector('.s-card__price') || card.querySelector('.s-item__price');
        if (priceEl) {
            const priceText = priceEl.textContent.trim();
            const pm = priceText.match(/([\d,]+\.?\d*)/);
            if (pm) {
                price = parseFloat(pm[1].replace(/,/g, ''));
                if (FORCE_CNY && price) price = Math.round(price / CNY_RATE * 100) / 100;
            }
        }

        let shipping = 0;
        let freeShipping = false;
        const footerEl = card.querySelector('.s-card__footer') || card.querySelector('.s-item__shipping, .s-item__logisticsCost');
        if (footerEl) {
            const footerText = footerEl.textContent.trim();
            const cleanText = footerText.replace(/[\u200b\u200c\u200d\ufeff\u2060\u2063]/g, '');
            if (cleanText.toLowerCase().includes('free') || cleanText.includes('$0.00')) {
                shipping = 0;
                freeShipping = true;
            } else {
                const sm = cleanText.match(/[¥￥$]?\s*([\d,]+\.?\d*)/);
                if (sm) {
                    shipping = parseFloat(sm[1].replace(/,/g, ''));
                    if (FORCE_CNY && shipping) shipping = Math.round(shipping / CNY_RATE * 100) / 100;
                }
            }
        }

        let imageUrl = '';
        const imgEl = card.querySelector('img.s-card__image') || card.querySelector('img.s-item__image-img');
        if (imgEl) {
            imageUrl = imgEl.getAttribute('src') || imgEl.getAttribute('data-src') || '';
            imageUrl = imageUrl.replace('/s-l140/', '/s-l300/').replace('/s-l200/', '/s-l300/');
        }

        let condition = '';
        const subEl = card.querySelector('.s-card__subtitle') || card.querySelector('.s-item__subtitle');
        if (subEl) {
            condition = subEl.textContent.trim();
        }

        let bestOffer = false;
        const allText = card.textContent.toLowerCase();
        if (allText.includes('best offer')) {
            bestOffer = true;
        }

        results.push({
            title, price, shipping, freeShipping, imageUrl,
            itemUrl, itemId, condition, bestOffer,
        });
    });

    return { items: results, totalResults };
}"""


def extract_items(page_obj) -> dict:
    """从 eBay 搜索结果页提取商品数据"""
    return page_obj.evaluate(JS_EXTRACT)


def save_to_excel(items: list, keyword: str, filepath: str):
    """将搜索结果保存为 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "eBay搜索结果"

    headers = ["序号", "标题", "价格($)", "运费($)", "总价($)", "免运费", "状况", "可议价", "商品ID", "商品URL", "图片URL", "来源页"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, item in enumerate(items, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=item.get('title', ''))
        ws.cell(row=row, column=3, value=item.get('price', 0)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=item.get('shipping', 0)).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=item.get('total', 0)).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value="✓" if item.get('free_shipping') else "").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=item.get('condition', ''))
        ws.cell(row=row, column=8, value="✓" if item.get('best_offer') else "").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=9, value=item.get('item_id', ''))
        ws.cell(row=row, column=10, value=item.get('url', ''))
        ws.cell(row=row, column=11, value=item.get('imageUrl', ''))
        ws.cell(row=row, column=12, value=item.get('page', '')).alignment = Alignment(horizontal="center")

    col_widths = [6, 60, 10, 10, 10, 8, 15, 8, 15, 45, 45, 8]
    for i, w in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"  Excel 已保存: {filepath}")


def scrape(keyword: str, sort: str = "best_match", last_n: int = 10,
           max_pages: int = 1, output_file: str = None, excel_file: str = None,
           progress_callback=None) -> dict:
    """爬取 eBay 搜索结果页面
    progress_callback(page_num, page_items): 每抓完一页调用，用于实时推送"""
    all_items = []
    ebay_total_results = 0

    # 获取实时 CNY → USD 汇率
    CNY_RATE = get_cny_usd_rate()
    print(f"[汇率] 当前 CNY/USD = {CNY_RATE:.4f}，人民币价格将自动转换为美元")

    try:
        stealth = Stealth()

        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-features=AutomationControlled',
                    '--no-sandbox',
                    '--disable-web-security',
                    '--disable-features=IsolateOrigins,site-per-process',
                ]
            )
            context = browser.new_context(
                user_agent=(
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                    '(KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36'
                ),
                locale='en-US',
                timezone_id='America/Los_Angeles',
                viewport={'width': 1920, 'height': 1080},
                screen={'width': 1920, 'height': 1080},
                extra_http_headers={
                    'Accept-Language': 'en-US,en;q=0.9',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                },
            )

            # ── 预设 eBay US 站点 cookie（语言/货币/地区） ──
            context.add_cookies([
                {
                    'name': 'dp1',
                    'value': 'b0b0',
                    'domain': '.ebay.com',
                    'path': '/',
                },
                {
                    'name': 'ebay',
                    'value': 'l%3Deng%26c%3DUS%26p%3DUS%26al%3DAs%20Is',
                    'domain': '.ebay.com',
                    'path': '/',
                },
                {
                    'name': 'nonsession',
                    'value': 'CgADLACBfOQJYiLjI5MjM3NDMzOGIZsZKKnJOAjYWIuMTJhYjBhMmZk',
                    'domain': '.ebay.com',
                    'path': '/',
                },
                # ← 关键：设定美国收货地址（邮编 90001 = Los Angeles, CA）
                {
                    'name': 'location',
                    'value': '{"zipCode":"90001","countryCode":"US","latitude":34.0522,"longitude":-118.2437,"radius":"25","unit":"MI"}',
                    'domain': '.ebay.com',
                    'path': '/',
                },
                # 备用：部分 eBay 版本用这个 cookie
                {
                    'name': 'shs',
                    'value': 's%3D90001',
                    'domain': '.ebay.com',
                    'path': '/',
                },
            ])
            stealth.apply_stealth_sync(context)
            page = context.new_page()

            # ── 通过 init_script 仅设定 navigator.languages（安全操作） ──
            # 注意：不在 init_script 中写 localStorage！
            # 原因：headless 模式下 init_script 执行时 localStorage 可能未就绪，
           #       抛出异常会阻塞 eBay 自身 JS 的执行，导致商品列表无法渲染。
            # localStorage 写入统一在 page.goto() 成功后通过 evaluate() 进行。
            print("  设定浏览器语言为 en-US ...")
            page.add_init_script("""
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['en-US', 'en'],
                });
            """)

            # ── 直接访问搜索页（跳过首页，避免重定向超时） ──
            print("  直接访问 eBay 搜索页...")
            first_url = build_url(keyword, sort=sort, page=1)
            resp = None
            for _att in range(1, 4):
                try:
                    resp = page.goto(first_url, wait_until='domcontentloaded', timeout=45000)
                    # 等待商品卡片实际渲染（eBay 用 JS 动态加载列表）
                    print("  等待商品列表加载...")
                    card_loaded = False
                    for _w in range(10):
                        try:
                            page.wait_for_selector('li.s-card, li.s-item', timeout=3000)
                            card_loaded = True
                            print(f"  商品列表已渲染 ({_w+1}s)")
                            break
                        except Exception:
                            page.wait_for_timeout(1000)
                    if not card_loaded:
                        print(f"  商品列表未渲染 (尝试 {_att}/3)")
                        if _att < 3:
                            page.wait_for_timeout(3000)
                            continue
                    # 检查是否被重定向到其他区域站点，如果是则用当前 URL 继续
                    final_url = page.url
                    if 'ebay.com' not in final_url or 'ebay.com.hk' in final_url or 'ebay.com.sg' in final_url:
                        print(f"  检测到重定向: {final_url}")
                        # 强制设置 cookie 回美国站
                        context.add_cookies([{
                            'name': 'dp1', 'value': 'b0b0',
                            'domain': '.ebay.com',
                            'path': '/'
                        }])
                        context.add_cookies([{
                            'name': 'ebay', 'value': 'l%3Deng%26c%3DUS%26p%3DUS',
                            'domain': '.ebay.com',
                            'path': '/'
                        }])
                        page.goto(first_url, wait_until='domcontentloaded', timeout=45000)
                        print("  已重新设置 US cookie 并重试")
                    # ── 先检查商品列表是否已正常渲染 ──
                    print("  检查商品列表渲染状态...")
                    initial_items = page.evaluate("""
                        () => {
                            const cards = document.querySelectorAll('li.s-card, li.s-item');
                            return cards.length;
                        }
                    """)
                    print(f"  当前商品卡片数: {initial_items}")

                    if initial_items > 0:
                        # 商品列表已经正常渲染，安全地注入地址信息
                        print("  商品列表已渲染，注入美国收货地址到 localStorage ...")
                        page.evaluate("""
                            try {
                                localStorage.setItem('dsrl', '90001');
                                localStorage.setItem('location', '{"zipCode":"90001","countryCode":"US","latitude":34.0522,"longitude":-118.2437,"radius":"25","unit":"MI"}');
                                localStorage.setItem('shippingCostRegion', 'US');
                            } catch(e) {}
                        """)
                        page.evaluate("""
                            (function() {
                                if (window.__SHIPPING_LOCATION__) return;
                                window.__SHIPPING_LOCATION__ = {
                                    postalCode: '90001',
                                    countryCode: 'US',
                                    latitude: 34.0522,
                                    longitude: -118.2437,
                                };
                            })()
                        """)
                        print("  地址注入完成 (Ship to: 90001 Los Angeles, CA)")
                    else:
                        # 商品列表未渲染，可能是 Cookie 中的 location 导致了问题
                        # 策略：先清除可疑 Cookie，重新加载页面（不加地址），能拿到结果总比拿不到强
                        print("  ⚠ 商品列表未渲染，可能地址 Cookie 导致异常")
                        print("  尝试移除 location Cookie 后重新加载...")
                        context.clear_cookies()
                        # 重新设置基础 Cookie（不含 location/shs）
                        context.add_cookies([
                            {
                                'name': 'dp1',
                                'value': 'b0b0',
                                'domain': '.ebay.com',
                                'path': '/',
                            },
                            {
                                'name': 'ebay',
                                'value': 'l%3Deng%26c%3DUS%26p%3DUS%26al%3DAs%20Is',
                                'domain': '.ebay.com',
                                'path': '/',
                            },
                        ])
                        stealth.apply_stealth_sync(context)  # 重新应用 stealth
                        page = context.new_page()
                        page.add_init_script("""
                            Object.defineProperty(navigator, 'languages', {
                                get: () => ['en-US', 'en'],
                            });
                        """)

                        resp = page.goto(first_url, wait_until='domcontentloaded', timeout=45000)
                        print("  重新加载完成（不含地址 Cookie）")

                        # 再次等待商品列表
                        card_loaded = False
                        for _w in range(10):
                            try:
                                page.wait_for_selector('li.s-card, li.s-item', timeout=3000)
                                card_loaded = True
                                print(f"  ✅ 商品列表已渲染 ({_w+1}s)")
                                break
                            except Exception:
                                page.wait_for_timeout(1000)

                        if not card_loaded:
                            raise Exception("移除地址 Cookie 后商品列表仍未渲染，eBay 可能检测到自动化行为")
                    break
                except Exception as _e:
                    print(f"  首次加载失败 (尝试 {_att}/3): {_e}")
                    if _att < 3:
                        page.wait_for_timeout(5000)
                    else:
                        raise

            # ── 翻页爬取 ──
            page_num = 0
            no_new_threshold = 0
            while True:
                page_num += 1

                if max_pages > 0 and page_num > max_pages:
                    print(f"  已达最大页数限制 ({max_pages})，停止翻页")
                    break

                url = build_url(keyword, sort=sort, page=page_num)
                print(f"  [p{page_num}] {url}")

                resp = None
                for _att in range(1, 4):
                    try:
                        resp = page.goto(url, wait_until='domcontentloaded', timeout=60000)
                        break
                    except Exception as _ge:
                        print(f"  页面加载超时 (尝试 {_att}/3): {_ge}")
                        if _att < 3:
                            page.wait_for_timeout(3000)
                        else:
                            print("  翻页失败，跳过此页")
                if resp is None:
                    no_new_threshold += 1
                    if no_new_threshold >= 3:
                        print("  连续失败过多，停止翻页")
                        break
                    continue
                status = resp.status if resp else 'N/A'
                print(f"  HTTP: {status}")

                if status == 403:
                    print("  403 拦截，等待重试...")
                    page.wait_for_timeout(5000)
                    resp = page.goto(url, wait_until='domcontentloaded', timeout=30000)
                    status = resp.status if resp else 'N/A'
                    print(f"  重试 HTTP: {status}")

                page.wait_for_timeout(random.randint(3000, 6000))

                try:
                    page.wait_for_selector('li.s-card, li.s-item', timeout=15000)
                except Exception:
                    debug_path = os.path.join(_SCRIPT_DIR, "_tmp_ebay_debug.html")
                    html = page.content()
                    with io.open(debug_path, 'w', encoding='utf-8') as f:
                        f.write(html[:200000])
                    print(f"  未找到商品卡片，页面已保存: {debug_path}")
                    break

                # 检测页面实际货币：设定美国地址后 eBay 应直接返回 USD
                # 如果页面已返回美元价格，则不需要 CNY→USD 转换
                page.evaluate(f"window.__CNY_RATE__ = {CNY_RATE};")
                currency_hint = page.evaluate("""
                    () => {
                        // 方法1：检查价格文本中是否包含 ¥ 或 CNY 前缀
                        const priceEl = document.querySelector('.s-card__price, .s-item__price');
                        if (priceEl) {
                            const txt = priceEl.textContent.trim();
                            if (txt.includes('¥') || txt.includes('CNY') || txt.includes('CN¥')) return 'CNY';
                        }
                        // 方法2：检查 URL 中的 currency 参数或页面上的货币标识
                        if (window.location.search.includes('currency=USD')) return 'USD';
                        // 默认认为是 USD（设了美国地址 cookie 后）
                        return 'USD';
                    }
                """)
                force_cny = (currency_hint == 'CNY')
                import json as _json
                page.evaluate(f"window.__FORCE_CNY__ = {_json.dumps(force_cny)};")
                if page_num == 1:
                    print(f"  货币检测: {currency_hint} | {'强制 CNY→USD 转换' if force_cny else '跳过转换（已是美元）'} (汇率 {CNY_RATE:.4f})")

                extract_data = extract_items(page)
                items_data = extract_data.get('items', [])
                page_total = extract_data.get('totalResults', 0)

                if page_num == 1 and page_total > 0:
                    ebay_total_results = page_total
                    max_page_estimate = (page_total + 59) // 60
                    print(f"  eBay 显示总结果数: {page_total}，预计最多 {max_page_estimate} 页")

                print(f"  本页提取 {len(items_data)} 个商品")

                if not items_data:
                    print("  无数据，停止翻页")
                    break

                if len(items_data) < 60:
                    print(f"  本页仅 {len(items_data)} 个商品，已到最后一页")
                    break
                else:
                    existing_ids = {it['item_id'] for it in all_items if it.get('item_id')}
                    new_in_page = [d for d in items_data if d.get('itemId') not in existing_ids]
                    if len(items_data) == 60 and len(new_in_page) == 0:
                        no_new_threshold += 1
                        print(f"  本页全部重复（连续 {no_new_threshold} 页无新商品）")
                        if no_new_threshold >= 2:
                            print("  连续2页无新商品，停止翻页")
                            break
                    else:
                        no_new_threshold = 0

                page_items = []
                for d in items_data:
                    p = d.get('price') or 0
                    s = d.get('shipping') if d.get('shipping') is not None else 0
                    total = p + s if p else 0
                    item = {
                        'title': d['title'],
                        'price': p,
                        'shipping': s,
                        'free_shipping': d.get('freeShipping', False),
                        'total': total,
                        'imageUrl': d.get('imageUrl', ''),
                        'url': d.get('itemUrl', ''),
                        'item_id': d.get('itemId', ''),
                        'condition': d.get('condition', ''),
                        'best_offer': d.get('bestOffer', False),
                        'page': page_num,
                    }
                    all_items.append(item)
                    page_items.append(item)
                    # 抓完一页立即打印该页结果
                    ship_str = "FREE" if item.get('free_shipping') else f"${s:.2f}"
                    cond = item.get('condition', '')[:15]
                    page_idx = len([x for x in all_items if x.get('page') == page_num])
                    print(f"  [p{page_num}|{page_idx:2d}] ${total:.2f} (${p:.2f}+{ship_str}) [{cond}] {item['title'][:55]}")

                print(f"  ✅ 第 {page_num} 页完成，本页 {len(items_data)} 个商品，已累计 {len(all_items)} 个")

                # 实时推送该页数据
                if progress_callback:
                    progress_callback(page_num, page_items)

                sleep_sec = random.uniform(2, 4)
                print(f"  等待 {sleep_sec:.1f}s 后翻页...")
                time.sleep(sleep_sec)

            browser.close()

    except Exception as e:
        import traceback
        import traceback as tb
        print(f"  异常: {e}")
        print(traceback.format_exc())
        err_msg = str(e).strip()
        if not err_msg:
            err_msg = f"爬取异常: {type(e).__name__}"
        return {"success": False, "error": err_msg}

    if not all_items:
        return {"success": False, "error": f"eBay 页面爬取未找到 '{keyword}' 的商品"}

    # 去重（按 item_id）
    seen_ids = set()
    unique_items = []
    dup_count = 0
    for item in all_items:
        iid = item.get('item_id', '')
        if iid and iid in seen_ids:
            dup_count += 1
            continue
        if iid:
            seen_ids.add(iid)
        unique_items.append(item)
    if dup_count:
        print(f"  去重: 移除 {dup_count} 个重复商品")

    # 保持原始顺序，不排序

    # 取最低N个算均价
    n = min(last_n, len(unique_items))
    lowest_n = unique_items[:n]
    avg = sum(x['total'] for x in lowest_n) / n

    result = {
        "success": True,
        "keyword": keyword,
        "sort": sort,
        "sort_desc": SORT_DESC.get(sort, sort),
        "ebay_total_results": ebay_total_results,
        "total_items": len(unique_items),
        "pages_scraped": page_num,
        "avg_price": round(avg, 2),
        "prices": [round(x['total'], 2) for x in lowest_n],
        "items": unique_items,
        "lowest_n": lowest_n,
    }

    # 保存 JSON
    if output_file:
        with io.open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
        print(f"  JSON 已保存: {output_file}")

    # 保存 Excel
    if excel_file:
        save_to_excel(unique_items, keyword, excel_file)

    return result


def scrape_data(keyword: str, last_n: int = 5, max_pages: int = 1,
                sort: str = "best_match", progress_callback=None, **kwargs) -> dict:
    """
    供 competitor_finder.py 调用的接口
    返回格式与 amazon_lowest_search.scrape_data 兼容
    """
    result = scrape(keyword, sort=sort, last_n=last_n, max_pages=max_pages,
                    progress_callback=progress_callback)

    if not result.get("success"):
        return result

    # 统一字段名，兼容 competitor_finder
    for item in result.get("items", []):
        item["image_url"] = item.get("imageUrl", "")
        item["shipping_cost"] = item.get("shipping", 0)

    result["source"] = "ebay_page"
    return result


# ═══════════════════ CLI 入口 ═══════════════════

if __name__ == "__main__":
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    parser = argparse.ArgumentParser(
        description="eBay 页面爬虫 - Playwright 浏览器搜索",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python ebay_page_scraper.py "Honda Pilot running boards"
  python ebay_page_scraper.py "关键词" --sort best_match --last 10 --pages 2
  python ebay_page_scraper.py "关键词" --pages 0 --excel

排序方式:
  best_match   最佳匹配（默认）
  price_low    价格：低到高
  price_high   价格：高到低
  price_ship   价格+运费：低到高
  newest       最新上架
  ending       即将结束
""",
    )

    parser.add_argument("keyword", help="搜索关键词")
    parser.add_argument("--sort", default="best_match",
                        choices=list(SORT_MAP.keys()),
                        help="排序方式（默认: best_match）")
    parser.add_argument("--last", type=int, default=10,
                        help="取最低N个算均价（默认10）")
    parser.add_argument("--pages", type=int, default=1,
                        help="最多翻页数（默认1，0=爬完全部页面）")
    parser.add_argument("--output", "-o", default=None,
                        help="输出JSON文件路径")
    parser.add_argument("--excel", action="store_true",
                        help="同时输出Excel文件")

    args = parser.parse_args()

    print("=" * 70)
    print(f"  eBay 页面爬虫 (Playwright)")
    print(f"  关键词: {args.keyword}")
    print(f"  排序: {SORT_DESC.get(args.sort, args.sort)}")
    print(f"  取最低 {args.last} 个 | 页数: {'全部' if args.pages == 0 else args.pages}")
    print("=" * 70)

    output = args.output or DEFAULT_OUTPUT
    excel_path = None
    if args.excel:
        base = os.path.splitext(output)[0]
        excel_path = base + ".xlsx"

    result = scrape(args.keyword, sort=args.sort, last_n=args.last,
                    max_pages=args.pages, output_file=output, excel_file=excel_path)

    if result.get("success"):
        items = result["items"]
        lowest = result.get("lowest_n", [])

        print()
        print("=" * 70)
        ebay_total = result.get('ebay_total_results', 0)
        total_str = f" / eBay总数: {ebay_total}" if ebay_total else ""
        print(f"  搜索完成！共爬取 {result.get('pages_scraped', '?')} 页，累计 {len(items)} 个商品{total_str}")
        print(f"  排序: {result.get('sort_desc', '')}")
        print("=" * 70)
        print(f"  >>> 均低价(最低{len(lowest)}个): ${result['avg_price']:.2f} <<<")
        print(f"  >>> 最低价: ${lowest[0].get('total', 0):.2f} - {lowest[0]['title'][:60]}")
    else:
        print(f"\n  失败: {result.get('error', '未知错误')}")
