"""
Amazon 搜索爬虫 - 自定义关键词 + 价格升序 + 最低N个均价
用法:
    python amazon_lowest_search.py "你的关键词"
    python amazon_lowest_search.py "关键词" --last 6 --pages 3
    python amazon_lowest_search.py "关键词" --pages 0 --excel
    python amazon_lowest_search.py "关键词" --output result.json --excel
"""
import re
import io
import sys
import json
import time
import random
import argparse
import os
from pathlib import Path
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

# 默认输出文件（相对于脚本所在目录）
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUTPUT = os.path.join(_SCRIPT_DIR, "_tmp_amazon_result.json")


def p(text='', out=None):
    if out:
        out.write(str(text) + '\n')
        out.flush()
    print(str(text))


def build_url(keyword, page=1, sort="relevance"):
    """根据关键词构建 Amazon 搜索URL（英文/美国站）
    sort: relevance(相关性,默认) / price_asc / price_desc / newest / rating
    """
    k = keyword.replace(' ', '+')
    first_word = keyword.split()[0].lower()
    sprefix = f"{first_word}%2Caps%2C344"
    qid = int(time.time())

    SORT_MAP = {
        "relevance": "relevanceblender",
        "price_asc": "price-asc-rank",
        "price_desc": "price-desc-rank",
        "newest": "date-desc-rank",
        "rating": "review-rank",
    }
    s_param = SORT_MAP.get(sort, "relevanceblender")

    url = (
        f"https://www.amazon.com/s?"
        f"k={k}"
        f"&s={s_param}"
        f"&language=en_US"
        f"&crid=2FGM3IOOBX1GW"
        f"&qid={qid}"
        f"&sprefix={sprefix}"
        f"&ref=sr_st_{s_param}"
    )
    if page > 1:
        url += f"&page={page}"
    return url


def extract_items(page_obj):
    """从当前页面提取商品数据，返回 list[dict]"""
    return page_obj.evaluate("""() => {
        const results = [];
        const divs = document.querySelectorAll('div[data-component-type="s-search-result"]');

        divs.forEach(div => {
            let title = '';
            const h2 = div.querySelector('h2');
            if (h2) {
                const span = h2.querySelector('span');
                title = span ? span.textContent.trim() : h2.textContent.trim();
            }
            if (!title) {
                const a = div.querySelector('a.a-link-normal');
                if (a) title = a.textContent.trim();
            }

            let price = null;
            // 兜底1: a-price-whole + a-price-fraction
            const priceWhole = div.querySelector('span.a-price-whole');
            const priceFrac = div.querySelector('span.a-price-fraction');
            if (priceWhole && priceFrac) {
                const pw = priceWhole.textContent.replace(/[^0-9]/g, '');
                const pf = priceFrac.textContent.replace(/[^0-9]/g, '');
                price = parseFloat(pw + '.' + pf);
            }
            // 兜底2: span.a-offscreen
            if (price === null || isNaN(price)) {
                const offscreen = div.querySelector('span.a-offscreen');
                if (offscreen) {
                    const m = offscreen.textContent.match(/([\\d,]+\\.?\\d*)/);
                    if (m) price = parseFloat(m[1].replace(/,/g, ''));
                }
            }
            // 兜底3: .a-price .a-offscreen 匹配 $xxx.xx
            if (price === null || isNaN(price)) {
                const priceEl = div.querySelector('.a-price .a-offscreen');
                if (priceEl) {
                    const m = priceEl.textContent.match(/\\$([\\d,]+\\.?\\d*)/);
                    if (m) price = parseFloat(m[1].replace(/,/g, ''));
                }
            }

            let freeShip = false;
            const text = div.textContent.toLowerCase();
            if (text.includes('free shipping') || text.includes('get free shipping')) {
                freeShip = true;
            }
            // 补充：CSS 选择器匹配免邮
            const freeShipEl = div.querySelector('[aria-label*="FREE Shipping"], [aria-label*="free shipping"]');
            if (freeShipEl) freeShip = true;

            let asin = div.getAttribute('data-asin') || '';

            // 提取商品图片URL
            let imageUrl = '';
            const imgEl = div.querySelector('img.s-image');
            if (imgEl) {
                imageUrl = imgEl.getAttribute('src') || imgEl.getAttribute('data-src') || '';
            }

            // 提取评分和评论数
            let rating = 0;
            let reviewCount = 0;
            const ratingEl = div.querySelector('span.a-icon-alt');
            if (ratingEl) {
                const rm = ratingEl.textContent.match(/([\\d.]+)\\s*out\\s*of\\s*5/);
                if (rm) rating = parseFloat(rm[1]);
            }
            const reviewEl = div.querySelector('span.a-size-base.s-underline-text');
            if (reviewEl) {
                const rcm = reviewEl.textContent.replace(/,/g, '').match(/(\\d+)/);
                if (rcm) reviewCount = parseInt(rcm[1]);
            }

            if (title && price && !isNaN(price)) {
                results.push({
                    title: title,
                    price: price,
                    freeShip: freeShip,
                    asin: asin,
                    imageUrl: imageUrl,
                    rating: rating,
                    reviewCount: reviewCount,
                });
            }
        });

        return results;
    }""")


def safe_goto(page, url, wait_until="domcontentloaded", timeout=45000, retries=3):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            resp = page.goto(url, wait_until=wait_until, timeout=timeout)
            page.wait_for_load_state("domcontentloaded", timeout=timeout)
            return resp
        except Exception as e:
            last_error = e
            if attempt >= retries:
                raise
            page.wait_for_timeout(1000 * attempt)
    raise last_error


def safe_extract_items(page, retries=2):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return extract_items(page)
        except Exception as e:
            last_error = e
            if "Execution context was destroyed" not in str(e) or attempt >= retries:
                raise
            page.wait_for_timeout(1200 * attempt)
            try:
                page.wait_for_load_state("domcontentloaded", timeout=15000)
            except Exception:
                pass
    raise last_error


def save_to_excel(items: list, keyword: str, filepath: str):
    """将搜索结果保存为 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon搜索结果"

    # 表头
    headers = ["序号", "标题", "价格($)", "免运费", "ASIN", "评分", "评论数", "商品URL", "图片URL", "来源页"]
    header_fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for i, item in enumerate(items, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=item.get('title', ''))
        ws.cell(row=row, column=3, value=item.get('price', 0)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value="✓" if item.get('free_ship') else "").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=item.get('asin', ''))
        ws.cell(row=row, column=6, value=item.get('rating', 0)).number_format = '0.0'
        ws.cell(row=row, column=7, value=item.get('reviewCount', 0))
        asin = item.get('asin', '')
        ws.cell(row=row, column=8, value=f"https://www.amazon.com/dp/{asin}" if asin else "")
        ws.cell(row=row, column=9, value=item.get('imageUrl', ''))
        ws.cell(row=row, column=10, value=item.get('page', '')).alignment = Alignment(horizontal="center")

    # 列宽
    col_widths = [6, 65, 10, 8, 15, 8, 8, 45, 45, 8]
    for i, w in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"  Excel 已保存: {filepath}")


def scrape(keyword, last_n=3, max_pages=1, output_file=None, excel_file=None, sort="relevance"):
    """爬取 Amazon 搜索结果，支持翻全页和 Excel 输出"""

    all_items = []
    amazon_total = 0

    try:
        stealth = Stealth()

        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=['--disable-blink-features=AutomationControlled']
            )
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                           '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
                locale='en-US',
                timezone_id='America/Los_Angeles',
                viewport={'width': 1920, 'height': 1080},
            )
            stealth.apply_stealth_sync(context)
            page = context.new_page()
            page.set_default_timeout(30000)
            page.set_default_navigation_timeout(45000)

            # 先访问首页，注入美国 locale cookie
            safe_goto(page, 'https://www.amazon.com', timeout=45000)
            page.wait_for_timeout(random.randint(1500, 3000))
            # 等待页面完全稳定（避免跳转导致 evaluate 上下文销毁）
            try:
                page.wait_for_load_state('networkidle', timeout=10000)
            except Exception:
                pass

            try:
                page.evaluate("""() => {
                    document.cookie = 'lc-main=enus; path=/; domain=.amazon.com';
                    document.cookie = 'i18n-prefs=USD; path=/; domain=.amazon.com';
                    document.cookie = 'sp-cdn=L5Z9CN; path=/; domain=.amazon.com';
                }""")
            except Exception as e:
                print(f"  [警告] cookie 注入失败（不影响搜索）: {e}")

            # 设置美国收货地址
            safe_goto(
                page,
                'https://www.amazon.com/gp/delivery/ajax/address-change.html'
                '?locationType=location&postalCode=90210&countryCode=US'
                '&city=Los+Angeles&state=CA&district=',
                timeout=30000
            )
            page.wait_for_timeout(random.randint(1000, 2000))
            print("  已设置收货地址: 美国 90210")

            safe_goto(page, 'https://www.amazon.com', timeout=45000)
            page.wait_for_timeout(random.randint(1000, 2000))

            # ── 翻页抓取 ──
            page_num = 0
            no_new_threshold = 0

            while True:
                page_num += 1

                if max_pages > 0 and page_num > max_pages:
                    print(f"  已达最大页数限制 ({max_pages})，停止翻页")
                    break

                url = build_url(keyword, page=page_num, sort=sort)
                print(f"  [p{page_num}] {url}")

                resp = safe_goto(page, url, timeout=45000)
                print(f"  HTTP: {resp.status if resp else 'N/A'}")

                # 等待页面稳定（防止二次跳转导致 evaluate 上下文销毁）
                try:
                    page.wait_for_load_state('networkidle', timeout=8000)
                except Exception:
                    pass  # 超时无妨，继续执行

                # 随机等待，模拟人类浏览行为
                wait_ms = random.randint(2500, 5000)
                page.wait_for_timeout(wait_ms)

                try:
                    page.wait_for_selector(
                        'div[data-component-type="s-search-result"]',
                        timeout=10000
                    )
                except Exception:
                    print(f"  第 {page_num} 页未找到商品，可能已到末页")
                    break

                # 提取总结果数（仅第一页）
                if page_num == 1:
                    try:
                        total_text = page.evaluate("""() => {
                            const el = document.querySelector('span[data-component-type="s-result-info-bar"]');
                            if (el) return el.textContent;
                            const h2 = document.querySelector('h2');
                            if (h2) return h2.textContent;
                            return '';
                        }""")
                        tm = re.search(r'([\d,]+)\s+result', total_text.replace(',', ''))
                        if tm:
                            amazon_total = int(tm.group(1).replace(',', ''))
                            print(f"  Amazon 显示总结果数: {amazon_total}")
                    except Exception as e:
                        print(f"  [警告] 提取总结果数失败: {e}")

                items_data = safe_extract_items(page)
                print(f"  本页提取 {len(items_data)} 个商品")

                if not items_data:
                    print("  本页无数据，停止翻页")
                    break

                # 检测重复
                existing_asins = {it['asin'] for it in all_items if it.get('asin')}
                new_in_page = [d for d in items_data if d.get('asin') not in existing_asins]

                if len(items_data) > 0 and len(new_in_page) == 0:
                    no_new_threshold += 1
                    print(f"  本页全部重复（连续 {no_new_threshold} 页无新商品）")
                    if no_new_threshold >= 2:
                        print("  连续2页无新商品，停止翻页")
                        break
                else:
                    no_new_threshold = 0

                for d in items_data:
                    all_items.append({
                        'title': d['title'],
                        'price': d['price'],
                        'ship': 0.0,
                        'free_ship': d['freeShip'],
                        'total': d['price'],
                        'asin': d.get('asin', ''),
                        'imageUrl': d.get('imageUrl', ''),
                        'rating': d.get('rating', 0),
                        'reviewCount': d.get('reviewCount', 0),
                        'page': page_num,
                    })

                # 检查是否有下一页按钮
                try:
                    has_next = page.evaluate("""() => {
                        const nextBtn = document.querySelector('a.s-pagination-next');
                        if (nextBtn && !nextBtn.classList.contains('s-pagination-item-disabled')) {
                            return true;
                        }
                        // 也检查 aria-disabled
                        const nextA = document.querySelector('a[aria-label="Next Page"]');
                        if (nextA && !nextA.hasAttribute('aria-disabled')) {
                            return true;
                        }
                        return false;
                    }""")
                except Exception as e:
                    print(f"  [警告] 检查下一页失败: {e}，停止翻页")
                    has_next = False

                if not has_next:
                    print(f"  没有下一页按钮，停止翻页")
                    break

                # 页间等待
                sleep_sec = random.uniform(2, 4)
                print(f"  等待 {sleep_sec:.1f}s 后翻页...")
                time.sleep(sleep_sec)

            browser.close()

    except Exception as e:
        import traceback
        print(f"  异常: {e}")
        print(traceback.format_exc())
        return {"success": False, "error": str(e)}

    if not all_items:
        return {"success": False, "error": f"Amazon 搜索未找到 '{keyword}' 的商品"}

    # 去重（按 asin）
    seen_asins = set()
    unique_items = []
    dup_count = 0
    for item in all_items:
        aid = item.get('asin', '')
        if aid and aid in seen_asins:
            dup_count += 1
            continue
        if aid:
            seen_asins.add(aid)
        unique_items.append(item)
    if dup_count:
        print(f"  去重: 移除 {dup_count} 个重复商品")

    # 按价格升序排序
    unique_items.sort(key=lambda x: x.get('total', 0) or 99999)

    # 取最低N个算均价
    n = min(last_n, len(unique_items))
    lowest_n = unique_items[:n]
    avg = sum(x['total'] for x in lowest_n) / n

    SORT_DESC_MAP = {
        "relevance": "Relevance",
        "price_asc": "Price: Low to High",
        "price_desc": "Price: High to Low",
        "newest": "Newest",
        "rating": "Avg. Customer Review",
    }
    result = {
        "success": True,
        "keyword": keyword,
        "sort": sort,
        "sort_desc": SORT_DESC_MAP.get(sort, sort),
        "amazon_total_results": amazon_total,
        "total_items": len(unique_items),
        "pages_scraped": page_num,
        "avg_price": round(avg, 2),
        "prices": [round(x['total'], 2) for x in lowest_n],
        "items": unique_items,
        "lowest_n": lowest_n,
    }

    # 保存 JSON
    if output_file:
        with io.open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
        print(f"  JSON 已保存: {output_file}")

    # 保存 Excel
    if excel_file:
        save_to_excel(unique_items, keyword, excel_file)

    return result


def scrape_data(keyword, last_n=3, max_pages=1, all_items=False, sort="relevance"):
    """
    搜索Amazon最低价商品，直接返回结构化数据
    供 competitor_finder.py 调用

    Args:
        all_items: 如果True，返回全部商品而非仅最低N个
        sort: 排序方式

    Returns: {
        "success": True/False,
        "avg_price": 平均价格,
        "prices": [价格列表],
        "items": [商品信息列表],
        "keyword": 关键词,
        "source": "amazon"
    }
    """
    result = scrape(keyword, last_n=last_n, max_pages=max_pages, sort=sort)

    if not result.get("success"):
        return result

    # 统一字段名，兼容 competitor_finder
    for item in result.get("items", []):
        item["image_url"] = item.get("imageUrl", "")
        item["url"] = f"https://www.amazon.com/dp/{item.get('asin', '')}" if item.get("asin") else ""
        item["shipping"] = 0
        item["total"] = item.get("price", 0)

    result["source"] = "amazon"

    # all_items 模式下返回全部商品
    if all_items:
        return result

    # 默认只返回最低N个
    n = min(last_n, len(result.get("items", [])))
    result["items"] = result.get("items", [])[:n]
    return result


if __name__ == '__main__':
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    parser = argparse.ArgumentParser(description='Amazon 搜索爬虫 - 最低价均价计算')
    parser.add_argument('keyword', help='搜索关键词')
    parser.add_argument('--last', type=int, default=3,
                        help='取最低N个算均价（默认3）')
    parser.add_argument('--pages', type=int, default=1,
                        help='最多抓取页数（默认1页，0=爬完全部页面）')
    parser.add_argument('--output', type=str, default=None,
                        help='输出JSON文件路径')
    parser.add_argument('--sort', default='relevance',
                        choices=['relevance', 'price_asc', 'price_desc', 'newest', 'rating'],
                        help='排序方式（默认: relevance）')
    parser.add_argument('--excel', action='store_true',
                        help='同时输出Excel文件（与JSON同目录）')

    args = parser.parse_args()

    print("=" * 70)
    print(f"  Amazon 搜索爬虫 (Playwright)")
    print(f"  关键词: {args.keyword}")
    print(f"  排序: {'Relevance' if args.sort == 'relevance' else args.sort}")
    print(f"  取最低 {args.last} 个 | 页数: {'全部' if args.pages == 0 else args.pages}")
    print("=" * 70)

    output = args.output or DEFAULT_OUTPUT
    excel_path = None
    if args.excel:
        base = os.path.splitext(output)[0]
        excel_path = base + ".xlsx"

    result = scrape(args.keyword, last_n=args.last, max_pages=args.pages,
                    output_file=output, excel_file=excel_path, sort=args.sort)

    if result.get("success"):
        items = result["items"]
        lowest = result.get("lowest_n", [])

        print()
        print("=" * 70)
        amz_total = result.get('amazon_total_results', 0)
        total_str = f" / Amazon总数: {amz_total}" if amz_total else ""
        print(f"  搜索结果: {len(items)} 个商品 (爬取 {result.get('pages_scraped', '?')} 页{total_str})")
        print(f"  排序: {result.get('sort_desc', '')}")
        print("=" * 70)

        for i, item in enumerate(items[:25], 1):
            ship_str = "FREE" if item.get('free_ship') else f"${item.get('ship', 0):.2f}"
            total = item.get('total', 0)
            asin = item.get('asin', '')
            print(f"  [{i:2d}] ${total:.2f} ({ship_str}) [{asin}] {item['title'][:50]}")

        if len(items) > 25:
            print(f"  ... 还有 {len(items) - 25} 条未显示")

        print()
        print("-" * 70)
        print(f"  最低 {len(lowest)} 个均价:")
        for i, item in enumerate(lowest, 1):
            print(f"    {i}. ${item.get('total', 0):.2f} - {item['title'][:60]}")
        print(f"  >>> 均低价: ${result['avg_price']:.2f} <<<")
        print("-" * 70)
    else:
        print(f"\n  失败: {result.get('error', '未知错误')}")
